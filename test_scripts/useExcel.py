import openpyxl
import xlrd
import xlwt
import glob
import sys
import os

""""
读取Excel文件需要的参数：
1、Excel文件的路径
2、指定数据的表名
3、指定单元格
"""

# excel_file_path = os.path.dirname(os.path.abspath(__file__))
# excel_file_names = glob.glob(excel_file_path + os.sep + "*.xls*")
# print(excel_file_path)
# print(excel_file_names)
# wb = openpyxl.load_workbook(excel_file_names[0])
# # sheet1为Excel文件的第一张表
# sheet1 = wb.worksheets[1]
# #sheet_names= wb.sheetnames
# # 打印第一张表的表名
# print(sheet1)
# #print(sheet_names)
# # 取第一张表中的A1单元格的文本
# # print(sheet1["A1"].value)

# for row in sheet1.iter_rows(min_col=1, max_col=3, min_row=1, max_row=1):
#     for cell in row:
#         print(cell.coordinate, cell.value)
#         print(type(cell.coordinate))


# 获取指定表的数据
def get_excel_data(file_path, sheet_name):
    # formatting_info=True --保持原样式
    work_book = openpyxl.load_workbook(file_path)
    # 获取所有的表名
    sheet_names = work_book.sheetnames
    print(sheet_names)
    print(sheet_names[1])
    sheet_data = work_book[sheet_name]
    print(sheet_data["A1"].value)


# 获取sheet表中指定行的数据
def get_excel_row_data(file_path, sheet_name, row_num):
    # formatting_info=True --保持原样式
    work_book = openpyxl.load_workbook(file_path)
    # 获取所有的表名
    sheet_names = work_book.sheetnames
    # 获取指定表名的数据
    sheet_data = work_book[sheet_name]
    # 获取指定行的数据
    sheet_row_data = sheet_data[row_num]
    sheet_row_data_values = []
    for cell in sheet_row_data:
        # print(cell.value)
        # 根据单元格的位置转化为具体的值
        sheet_row_data_values.append(cell.value)
    print(sheet_row_data_values)


# 获取sheet表中指定列的数据
def get_excel_col_data(file_path, sheet_name, col_num):
    # formatting_info=True --保持原样式
    work_book = openpyxl.load_workbook(file_path)
    # 获取所有的表名
    sheet_names = work_book.sheetnames
    # 获取指定表名的数据
    sheet_data = work_book[sheet_name]
    # 获取表的行数和列数
    rows = sheet_data.max_row
    cols = sheet_data.max_column
    print("表格行数为：", rows)
    print("表格列数为：", cols)
    # 获取指定列的数据
    sheet_col_data = sheet_data[col_num]
    sheet_col_data_values = []
    for cell in sheet_col_data:
        # 判断一下列数据是不是空的，如果是空的则过滤掉
        if cell.value is not None:
            # 根据单元格的位置转化为具体的值
            sheet_col_data_values.append(cell.value)
    print(sheet_col_data_values)


if __name__ == "__main__":
    # get_excel_data('../test_data/testexcel.xlsx', 'parking_zx')
    get_excel_row_data('../test_data/testexcel.xlsx', 'Sheet1', 2)
    # get_excel_col_data('../test_data/testexcel.xlsx', 'Sheet1', "B")
